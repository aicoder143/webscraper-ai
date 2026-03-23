import csv
import json
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def export_json(site, pages, analysis=None) -> str:
    data = {
        "export_meta": {
            "generated_at": datetime.now().isoformat(),
            "tool": "WebScraper AI",
            "version": "1.0",
        },
        "site": {
            "id":          site.id,
            "url":         site.url,
            "status":      site.status,
            "scrape_mode": site.scrape_mode,
            "scraped_at":  site.created_at.isoformat(),
        },
        "stats": {
            "total_pages":        len(pages),
            "pages_with_content": sum(1 for p in pages if p.word_count > 0),
            "total_words":        sum(p.word_count for p in pages),
        },
        "pages": [
            {
                "url":              p.page_url,
                "title":            p.title,
                "word_count":       p.word_count,
                "meta_description": p.meta_description,
                "headings":         p.headings,
                "content_preview":  p.content[:500] if p.content else "",
            }
            for p in pages
        ],
        "analysis": analysis.extracted_data if analysis else None,
    }
    return json.dumps(data, indent=2, ensure_ascii=False)


def export_csv_pages(pages) -> str:
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow([
        "URL", "Title", "Word Count",
        "Meta Description", "H1 Headings",
        "H2 Headings", "Content Preview",
    ])
    for p in pages:
        h1s = " | ".join(
            h["text"] for h in (p.headings or [])
            if h.get("level") == "h1"
        )
        h2s = " | ".join(
            h["text"] for h in (p.headings or [])
            if h.get("level") == "h2"
        )
        writer.writerow([
            p.page_url,
            p.title or "",
            p.word_count,
            p.meta_description or "",
            h1s,
            h2s,
            (p.content or "")[:500].replace("\n", " "),
        ])
    return output.getvalue()


def export_csv_analysis(site, analysis) -> str:
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(["Field", "Value"])
    writer.writerow(["Site URL", site.url])
    writer.writerow(["Scraped At", site.created_at.isoformat()])
    writer.writerow(["Scrape Mode", site.scrape_mode or ""])
    writer.writerow(["", ""])
    if analysis:
        d = analysis.extracted_data
        fields = [
            ("Business Name",     "business_name"),
            ("Business Type",     "business_type"),
            ("Description",       "description"),
            ("Contact Email",     "contact_email"),
            ("Contact Phone",     "contact_phone"),
            ("Location",          "location"),
            ("Language",          "language"),
            ("Sentiment",         "sentiment"),
            ("Extraction Method", "extraction_method"),
            ("Pages Analyzed",    "total_pages_analyzed"),
        ]
        for label, key in fields:
            writer.writerow([label, d.get(key) or ""])
        writer.writerow(["Main Topics",
            " | ".join(d.get("main_topics", []))])
        writer.writerow(["Products/Services",
            " | ".join(d.get("products_services", []))])
        writer.writerow(["Social Links",
            " | ".join(d.get("social_links", []))])
        writer.writerow(["Key Facts",
            " | ".join(d.get("key_facts", []))])
    return output.getvalue()


def export_excel(site, pages, analysis=None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb        = openpyxl.Workbook()
    DARK      = PatternFill("solid", fgColor="0D1117")
    GRAY      = PatternFill("solid", fgColor="F0F4F8")
    WHITE     = PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    BODY_FONT = Font(color="2C3E50", size=10)
    thin      = Side(style="thin", color="D0DCE8")
    BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, cols):
        for c in range(1, cols + 1):
            cell            = ws.cell(row=row, column=c)
            cell.fill       = DARK
            cell.font       = HDR_FONT
            cell.alignment  = Alignment(horizontal="center", vertical="center")
            cell.border     = BRD

    def row_style(ws, row, cols, alt=False):
        for c in range(1, cols + 1):
            cell            = ws.cell(row=row, column=c)
            cell.fill       = GRAY if alt else WHITE
            cell.font       = BODY_FONT
            cell.alignment  = Alignment(vertical="top", wrap_text=True)
            cell.border     = BRD

    # Sheet 1 Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 55
    total_words   = sum(p.word_count for p in pages)
    pages_content = sum(1 for p in pages if p.word_count > 0)
    summary = [
        ("WebScraper AI Export", ""),
        ("", ""),
        ("Site URL",     site.url),
        ("Scraped At",   site.created_at.strftime("%Y-%m-%d %H:%M")),
        ("Scrape Mode",  site.scrape_mode or "http"),
        ("Status",       site.status),
        ("", ""),
        ("Total Pages",  len(pages)),
        ("With Content", pages_content),
        ("Total Words",  f"{total_words:,}"),
        ("PDF",          "Yes" if site.pdf_file else "No"),
        ("AI Analysis",  "Yes" if analysis else "No"),
    ]
    for i, (k, v) in enumerate(summary, 1):
        ws1.cell(row=i, column=1, value=k).font = Font(
            bold=True, size=11 if i == 1 else 10,
            color="0099BB" if i == 1 else "2C3E50"
        )
        ws1.cell(row=i, column=2, value=str(v)).font = BODY_FONT

    # Sheet 2 Pages
    ws2 = wb.create_sheet("Pages")
    headers = ["URL","Title","Words","Meta","H1","H2","Content Preview"]
    widths  = [45, 35, 8, 40, 35, 35, 60]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws2.cell(row=1, column=i, value=h)
        ws2.column_dimensions[get_column_letter(i)].width = w
    hdr(ws2, 1, len(headers))
    ws2.row_dimensions[1].height = 22
    for r, p in enumerate(pages, 2):
        h1s = " | ".join(h["text"] for h in (p.headings or []) if h.get("level") == "h1")
        h2s = " | ".join(h["text"] for h in (p.headings or []) if h.get("level") == "h2")
        for c, val in enumerate([
            p.page_url, p.title or "", p.word_count,
            p.meta_description or "", h1s, h2s,
            (p.content or "")[:400].replace("\n", " ")
        ], 1):
            ws2.cell(row=r, column=c, value=val)
        row_style(ws2, r, len(headers), alt=(r % 2 == 0))
        ws2.row_dimensions[r].height = 18
    ws2.freeze_panes = "A2"

    # Sheet 3 Analysis
    if analysis:
        ws3 = wb.create_sheet("AI Analysis")
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 65
        ws3.cell(row=1, column=1, value="Field")
        ws3.cell(row=1, column=2, value="Value")
        hdr(ws3, 1, 2)
        d = analysis.extracted_data
        an = [
            ("Business Name",     d.get("business_name","") or ""),
            ("Business Type",     d.get("business_type","") or ""),
            ("Description",       d.get("description","") or ""),
            ("Contact Email",     d.get("contact_email","") or ""),
            ("Contact Phone",     d.get("contact_phone","") or ""),
            ("Location",          d.get("location","") or ""),
            ("Language",          d.get("language","") or ""),
            ("Sentiment",         d.get("sentiment","") or ""),
            ("Main Topics",       " | ".join(d.get("main_topics",[]))),
            ("Products/Services", " | ".join(d.get("products_services",[]))),
            ("Social Links",      " | ".join(d.get("social_links",[]))),
            ("Key Facts",         " | ".join(d.get("key_facts",[]))),
            ("Method",            d.get("extraction_method","") or ""),
            ("Pages Analyzed",    str(d.get("total_pages_analyzed",""))),
        ]
        for r, (k, v) in enumerate(an, 2):
            ws3.cell(row=r, column=1, value=k).font = Font(bold=True, color="2C3E50", size=10)
            ws3.cell(row=r, column=2, value=v).font = BODY_FONT
            row_style(ws3, r, 2, alt=(r % 2 == 0))
            ws3.row_dimensions[r].height = 18
        ws3.freeze_panes = "A2"

    # Sheet 4 Content Index
    ws4 = wb.create_sheet("Content Index")
    ws4.column_dimensions["A"].width = 45
    ws4.column_dimensions["B"].width = 35
    ws4.column_dimensions["C"].width = 10
    for c, h in enumerate(["URL","Title","Words"], 1):
        ws4.cell(row=1, column=c, value=h)
    hdr(ws4, 1, 3)
    for r, p in enumerate(
        sorted(pages, key=lambda x: x.word_count, reverse=True), 2
    ):
        ws4.cell(row=r, column=1, value=p.page_url)
        ws4.cell(row=r, column=2, value=p.title or "")
        ws4.cell(row=r, column=3, value=p.word_count)
        row_style(ws4, r, 3, alt=(r % 2 == 0))
    ws4.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
